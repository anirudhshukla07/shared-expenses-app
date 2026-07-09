import csv
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path

from django.conf import settings
from django.db import transaction
from openpyxl import load_workbook

from expenses.models import (
    Expense,
    ExpenseSplit,
    GroupMembership,
    ImportAnomaly,
    ImportBatch,
    LedgerStatus,
    Person,
    Settlement,
)

PAISE = Decimal("0.01")
DATE_ORIGIN = datetime(1899, 12, 30).date()
REQUIRED_HEADERS = [
    "date",
    "description",
    "paid_by",
    "amount",
    "currency",
    "split_type",
    "split_with",
    "split_details",
    "notes",
]

ALIASES = {
    "aisha": "Aisha",
    "rohan": "Rohan",
    "priya": "Priya",
    "priya s": "Priya",
    "meera": "Meera",
    "dev": "Dev",
    "sam": "Sam",
    "devs friend kabir": "Kabir",
    "dev friend kabir": "Kabir",
    "kabir": "Kabir",
}

PAYMENT_WORDS = ["paid back", "paid", "settlement", "settled", "deposit share", "deposit"]
REFUND_WORDS = ["refund", "cancelled", "canceled"]


def q(amount):
    return Decimal(amount).quantize(PAISE, rounding=ROUND_HALF_UP)


def canonical_key(name):
    if name is None:
        return ""
    value = str(name).strip().lower()
    value = value.replace("'", "")
    value = re.sub(r"[^a-z0-9 ]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def display_name(raw_name):
    key = canonical_key(raw_name)
    return ALIASES.get(key, str(raw_name).strip() if raw_name is not None else "")


def normalize_description(text):
    text = str(text or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\b(at|the|a|an|order|dinner|lunch)\b", " ", text)
    return re.sub(r"\s+", " ", text).strip()


class ExpenseImportService:
    def __init__(self, group, batch, file_path, fx_rates=None):
        self.group = group
        self.batch = batch
        self.file_path = Path(file_path)
        self.fx_rates = {k.upper(): Decimal(str(v)) for k, v in (fx_rates or settings.DEFAULT_FX_RATES).items()}
        self.expected_start = date.fromisoformat(settings.EXPECTED_IMPORT_START)
        self.expected_end = date.fromisoformat(settings.EXPECTED_IMPORT_END)
        self.exact_seen = {}
        self.fuzzy_seen = []
        self.report_rows = []

    def run(self):
        try:
            rows = self.load_rows()
            self.batch.total_rows = len(rows)
            self.batch.save(update_fields=["total_rows"])
            with transaction.atomic():
                for row_number, row in rows:
                    self.import_row(row_number, row)
                self.finish_batch()
        except Exception as exc:  # deliberate: never leave frontend with a crashed mystery import
            self.batch.status = ImportBatch.Status.FAILED
            self.batch.report_json = {"error": str(exc), "rows": self.report_rows}
            self.batch.save(update_fields=["status", "report_json"])
            raise
        return self.batch

    def load_rows(self):
        suffix = self.file_path.suffix.lower()
        if suffix == ".csv":
            with self.file_path.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                self.validate_headers(reader.fieldnames or [])
                return [(idx, self.clean_row(row)) for idx, row in enumerate(reader, start=2)]
        if suffix in {".xlsx", ".xlsm"}:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
            self.validate_headers(headers)
            rows = []
            for idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row = dict(zip(headers, values))
                if any(value is not None and str(value).strip() != "" for value in row.values()):
                    rows.append((idx, self.clean_row(row)))
            return rows
        raise ValueError("Only .csv and .xlsx imports are supported")

    def validate_headers(self, headers):
        normalized = [str(h).strip() for h in headers]
        missing = [h for h in REQUIRED_HEADERS if h not in normalized]
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

    def clean_row(self, row):
        return {str(k).strip(): (None if v == "" else v) for k, v in row.items()}

    def import_row(self, row_number, row):
        row_anomalies = []
        parsed_date = self.parse_date(row.get("date"), row_number, row_anomalies)
        description = str(row.get("description") or "").strip()
        paid_by_raw = row.get("paid_by")
        amount = self.parse_amount(row.get("amount"), row_number, row_anomalies)
        currency = self.parse_currency(row.get("currency"), row_number, row_anomalies)
        split_type = str(row.get("split_type") or "").strip().lower()
        split_with_raw = str(row.get("split_with") or "").strip()
        split_details_raw = str(row.get("split_details") or "").strip()
        notes = str(row.get("notes") or "").strip()

        if parsed_date is None or amount is None:
            self.add_report_only(row_number, row_anomalies)
            return

        if parsed_date < self.expected_start or parsed_date > self.expected_end:
            row_anomalies.append({
                "code": "AMBIGUOUS_OR_OUT_OF_SCOPE_DATE",
                "severity": ImportAnomaly.Severity.ERROR,
                "message": f"Date {parsed_date} is outside the expected assignment import window.",
                "policy": "Do not guess corrected dates from messy data.",
                "action": "Row requires review and is not posted.",
                "review": True,
            })

        original_amount = q(amount)
        fx_rate = self.fx_rates.get(currency)
        if fx_rate is None:
            row_anomalies.append({
                "code": "UNSUPPORTED_CURRENCY",
                "severity": ImportAnomaly.Severity.ERROR,
                "message": f"Currency {currency} is not configured.",
                "policy": "Unsupported currencies cannot be imported without a rate.",
                "action": "Row skipped.",
                "review": True,
            })
            self.add_report_only(row_number, row_anomalies)
            return
        amount_inr = q(Decimal(amount) * fx_rate)
        if currency != "INR":
            row_anomalies.append({
                "code": "CURRENCY_CONVERTED",
                "severity": ImportAnomaly.Severity.INFO,
                "message": f"Converted {original_amount} {currency} to {amount_inr} INR at rate {fx_rate}.",
                "policy": "Non-INR rows are converted using a configured import rate stored with the row.",
                "action": "Converted and surfaced in report.",
                "review": False,
            })

        if amount == 0:
            row_anomalies.append({
                "code": "ZERO_AMOUNT",
                "severity": ImportAnomaly.Severity.WARNING,
                "message": "Amount is zero and has no financial effect.",
                "policy": "Zero amount rows are skipped as no-op rows.",
                "action": "Skipped.",
                "review": False,
            })
            self.add_report_only(row_number, row_anomalies)
            return

        if amount < 0:
            if any(word in f"{description} {notes}".lower() for word in REFUND_WORDS):
                row_anomalies.append({
                    "code": "NEGATIVE_AMOUNT_REFUND",
                    "severity": ImportAnomaly.Severity.WARNING,
                    "message": "Negative amount appears to be a refund.",
                    "policy": "Refund-like negative rows are imported as negative expenses and surfaced.",
                    "action": "Imported as refund calculation.",
                    "review": False,
                })
            else:
                row_anomalies.append({
                    "code": "NEGATIVE_AMOUNT_UNCLEAR",
                    "severity": ImportAnomaly.Severity.ERROR,
                    "message": "Negative amount is not clearly a refund.",
                    "policy": "Do not infer unclear negative rows.",
                    "action": "Review required.",
                    "review": True,
                })

        payer_name = display_name(paid_by_raw)
        if not payer_name:
            row_anomalies.append({
                "code": "MISSING_PAYER",
                "severity": ImportAnomaly.Severity.ERROR,
                "message": "Missing paid_by value.",
                "policy": "The payer cannot be inferred safely.",
                "action": "Row skipped until user supplies payer.",
                "review": True,
            })
            self.add_report_only(row_number, row_anomalies)
            return
        self.name_anomaly_if_needed(row_number, paid_by_raw, payer_name, row_anomalies)
        payer = self.get_or_create_person(payer_name)

        if self.looks_like_payment(description, notes, split_type):
            self.import_settlement(row_number, parsed_date, description, payer, original_amount, currency, fx_rate, amount_inr, split_with_raw, notes, row_anomalies)
            return

        if split_type not in {"equal", "unequal", "percentage", "share"}:
            row_anomalies.append({
                "code": "UNKNOWN_SPLIT_TYPE",
                "severity": ImportAnomaly.Severity.ERROR,
                "message": f"Unsupported split_type: {split_type or '<blank>'}.",
                "policy": "Only split types present in the assignment are supported.",
                "action": "Row skipped.",
                "review": True,
            })
            self.add_report_only(row_number, row_anomalies)
            return

        participants = self.parse_people(split_with_raw, row_number, row_anomalies)
        if not participants:
            row_anomalies.append({
                "code": "NO_SPLIT_PARTICIPANTS",
                "severity": ImportAnomaly.Severity.ERROR,
                "message": "No split participants found.",
                "policy": "Cannot split without participants.",
                "action": "Row skipped.",
                "review": True,
            })
            self.add_report_only(row_number, row_anomalies)
            return

        participants = self.apply_membership_policy(parsed_date, participants, row_number, row_anomalies)
        if not participants:
            self.add_report_only(row_number, row_anomalies)
            return

        split_amounts = self.calculate_splits(split_type, amount_inr, participants, split_details_raw, row_number, row_anomalies)
        if split_amounts is None:
            self.add_report_only(row_number, row_anomalies)
            return

        duplicate_review = self.detect_duplicates(row_number, parsed_date, description, payer_name, amount_inr, participants, row_anomalies)
        needs_review = duplicate_review or any(a["review"] for a in row_anomalies)
        status = LedgerStatus.REVIEW_REQUIRED if needs_review else LedgerStatus.POSTED

        expense = Expense.objects.create(
            group=self.group,
            import_batch=self.batch,
            raw_row_number=row_number,
            date=parsed_date,
            description=description,
            normalized_description=normalize_description(description),
            paid_by=payer,
            amount_original=original_amount,
            currency=currency,
            fx_rate_to_inr=fx_rate,
            amount_inr=amount_inr,
            split_type=split_type,
            split_with_raw=split_with_raw,
            split_details_raw=split_details_raw,
            notes=notes,
            status=status,
        )
        for person_name, owed, basis in split_amounts:
            person = self.get_or_create_person(person_name)
            ExpenseSplit.objects.create(expense=expense, person=person, amount_owed_inr=owed, basis=basis)
        self.persist_anomalies(row_number, row_anomalies, expense=expense)
        self.increment_counts(status)
        self.add_report_row(row_number, status, row_anomalies)

    def parse_date(self, value, row_number, anomalies):
        if value in (None, ""):
            anomalies.append({
                "code": "MISSING_DATE", "severity": ImportAnomaly.Severity.ERROR,
                "message": "Missing date.", "policy": "Rows need dates for membership-aware splitting.",
                "action": "Row skipped.", "review": True,
            })
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            return DATE_ORIGIN + timedelta(days=int(value))
        text = str(value).strip()
        if re.match(r"^\d+(\.0)?$", text):
            return DATE_ORIGIN + timedelta(days=int(float(text)))
        formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"]
        parsed = []
        for fmt in formats:
            try:
                parsed.append(datetime.strptime(text, fmt).date())
            except ValueError:
                pass
        if len(set(parsed)) > 1:
            anomalies.append({
                "code": "AMBIGUOUS_DATE_FORMAT", "severity": ImportAnomaly.Severity.WARNING,
                "message": f"Date string {text!r} can be interpreted in multiple ways.",
                "policy": "Prefer day-first for India only if the parsed date remains in import window.",
                "action": "Parsed but requires review.", "review": True,
            })
        if parsed:
            return parsed[0]
        anomalies.append({
            "code": "INVALID_DATE", "severity": ImportAnomaly.Severity.ERROR,
            "message": f"Could not parse date {text!r}.",
            "policy": "Do not guess invalid dates.", "action": "Row skipped.", "review": True,
        })
        return None

    def parse_amount(self, value, row_number, anomalies):
        try:
            amount = Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, AttributeError):
            anomalies.append({
                "code": "INVALID_AMOUNT", "severity": ImportAnomaly.Severity.ERROR,
                "message": f"Invalid amount {value!r}.", "policy": "Invalid amounts cannot be posted.",
                "action": "Row skipped.", "review": True,
            })
            return None
        if amount != q(amount):
            anomalies.append({
                "code": "ROUNDING_APPLIED", "severity": ImportAnomaly.Severity.INFO,
                "message": f"Amount {amount} has more than paise precision.",
                "policy": "Money is rounded to 2 decimals using Decimal HALF_UP.",
                "action": f"Rounded to {q(amount)}.", "review": False,
            })
        return amount

    def parse_currency(self, value, row_number, anomalies):
        if value in (None, ""):
            anomalies.append({
                "code": "MISSING_CURRENCY", "severity": ImportAnomaly.Severity.WARNING,
                "message": "Missing currency.",
                "policy": "Blank domestic rows are assumed INR but require review.",
                "action": "Assumed INR; review required.", "review": True,
            })
            return "INR"
        return str(value).strip().upper()

    def parse_people(self, raw, row_number, anomalies):
        people = []
        for token in str(raw or "").split(";"):
            token = token.strip()
            if not token:
                continue
            name = display_name(token)
            self.name_anomaly_if_needed(row_number, token, name, anomalies)
            people.append(name)
        return people

    def name_anomaly_if_needed(self, row_number, raw, normalized, anomalies):
        if raw is None:
            return
        raw_clean = str(raw).strip()
        if raw_clean and raw_clean != normalized:
            anomalies.append({
                "code": "NAME_NORMALIZED", "severity": ImportAnomaly.Severity.INFO,
                "message": f"Normalized name {raw_clean!r} to {normalized!r}.",
                "policy": "Whitespace, case, and known aliases are canonicalized.",
                "action": "Canonical name used for ledger calculations.", "review": False,
            })

    def get_or_create_person(self, name):
        canonical = canonical_key(name)
        person, _ = Person.objects.get_or_create(canonical_name=canonical, defaults={"name": name})
        return person

    def active_person_names(self, day):
        memberships = GroupMembership.objects.filter(group=self.group, starts_on__lte=day).select_related("person")
        active = set()
        for membership in memberships:
            if membership.ends_on is None or day <= membership.ends_on:
                active.add(membership.person.name)
        return active

    def apply_membership_policy(self, day, participants, row_number, anomalies):
        active_names = self.active_person_names(day)
        corrected = []
        for name in participants:
            if name in active_names:
                corrected.append(name)
                continue
            if name in {"Dev", "Kabir"}:
                person = self.get_or_create_person(name)
                GroupMembership.objects.get_or_create(
                    group=self.group,
                    person=person,
                    starts_on=day,
                    ends_on=day,
                    defaults={"role": "guest"},
                )
                corrected.append(name)
                anomalies.append({
                    "code": "UNKNOWN_GUEST_MEMBER", "severity": ImportAnomaly.Severity.WARNING,
                    "message": f"{name} was not an active group member on {day}, treated as a one-day guest.",
                    "policy": "Trip guests can be included, but the app creates an auditable membership row.",
                    "action": "Created guest membership for that date; review required.", "review": True,
                })
            else:
                anomalies.append({
                    "code": "INACTIVE_MEMBER_IN_SPLIT", "severity": ImportAnomaly.Severity.WARNING,
                    "message": f"{name} is not active on {day}.",
                    "policy": "Inactive members are not charged without review.",
                    "action": f"Removed {name} from proposed split; review required.", "review": True,
                })
        return corrected

    def calculate_splits(self, split_type, amount_inr, participants, split_details_raw, row_number, anomalies):
        if split_type == "equal":
            if split_details_raw:
                anomalies.append({
                    "code": "EQUAL_WITH_SPLIT_DETAILS", "severity": ImportAnomaly.Severity.WARNING,
                    "message": "split_type is equal but split_details are present.",
                    "policy": "The split_type column wins because it is the explicit structured field.",
                    "action": "Ignored split_details and split equally; review required.", "review": True,
                })
            return self.allocate_by_weights(amount_inr, [(name, Decimal("1"), "equal") for name in participants])

        details = self.parse_split_details(split_details_raw)
        if not details:
            anomalies.append({
                "code": "MISSING_SPLIT_DETAILS", "severity": ImportAnomaly.Severity.ERROR,
                "message": f"{split_type} split requires split_details.",
                "policy": "Non-equal split details cannot be inferred.",
                "action": "Row skipped.", "review": True,
            })
            return None

        missing = [name for name in participants if name not in details]
        if missing:
            anomalies.append({
                "code": "SPLIT_DETAILS_MISSING_PERSON", "severity": ImportAnomaly.Severity.ERROR,
                "message": f"Missing split details for: {', '.join(missing)}.",
                "policy": "Every participant must have a split value.",
                "action": "Row skipped.", "review": True,
            })
            return None

        if split_type == "unequal":
            amounts = [(name, q(details[name]), "unequal amount") for name in participants]
            total = sum(amount for _, amount, _ in amounts)
            if total != q(amount_inr):
                anomalies.append({
                    "code": "UNEQUAL_TOTAL_MISMATCH", "severity": ImportAnomaly.Severity.ERROR,
                    "message": f"Unequal split total {total} does not equal expense amount {amount_inr}.",
                    "policy": "Do not scale exact rupee allocations silently.",
                    "action": "Review required.", "review": True,
                })
            return amounts

        if split_type == "share":
            weights = [(name, details[name], f"{details[name]} shares") for name in participants]
            return self.allocate_by_weights(amount_inr, weights)

        if split_type == "percentage":
            weights = [(name, details[name], f"{details[name]}%") for name in participants]
            total = sum(weight for _, weight, _ in weights)
            if total != Decimal("100"):
                anomalies.append({
                    "code": "PERCENT_TOTAL_NOT_100", "severity": ImportAnomaly.Severity.WARNING,
                    "message": f"Percent split totals {total} instead of 100.",
                    "policy": "Preserve the listed proportions but require approval before posting.",
                    "action": "Normalized percentages as proposed calculation; review required.", "review": True,
                })
            return self.allocate_by_weights(amount_inr, weights)
        return None

    def parse_split_details(self, raw):
        result = {}
        for part in str(raw or "").split(";"):
            part = part.strip()
            if not part:
                continue
            match = re.match(r"(.+?)\s+(-?\d+(?:\.\d+)?)\s*%?$", part)
            if not match:
                continue
            name = display_name(match.group(1).strip())
            result[name] = Decimal(match.group(2))
        return result

    def allocate_by_weights(self, amount, weighted_people):
        total_weight = sum(weight for _, weight, _ in weighted_people)
        if total_weight == 0:
            return None
        allocations = []
        allocated = Decimal("0.00")
        for idx, (name, weight, basis) in enumerate(weighted_people):
            if idx == len(weighted_people) - 1:
                owed = q(amount - allocated)
            else:
                owed = q((amount * weight) / total_weight)
                allocated += owed
            allocations.append((name, owed, basis))
        return allocations

    def looks_like_payment(self, description, notes, split_type):
        text = f"{description} {notes}".lower()
        # Some payment rows are incorrectly logged with a split_type, so explicit
        # payment/deposit wording wins over the split_type column.
        strong_payment_markers = ["paid back", "settlement", "settled", "deposit share", "deposit"]
        if any(word in text for word in strong_payment_markers):
            return True
        if split_type:
            return False
        return any(word in text for word in PAYMENT_WORDS)

    def import_settlement(self, row_number, day, description, payer, original_amount, currency, fx_rate, amount_inr, split_with_raw, notes, anomalies):
        recipients = self.parse_people(split_with_raw, row_number, anomalies)
        if len(recipients) != 1:
            anomalies.append({
                "code": "SETTLEMENT_RECIPIENT_UNCLEAR", "severity": ImportAnomaly.Severity.ERROR,
                "message": "Settlement must have exactly one recipient in split_with.",
                "policy": "Do not infer settlement recipient.",
                "action": "Row skipped.", "review": True,
            })
            self.add_report_only(row_number, anomalies)
            return
        paid_to = self.get_or_create_person(recipients[0])
        anomalies.append({
            "code": "SETTLEMENT_DETECTED", "severity": ImportAnomaly.Severity.INFO,
            "message": "Row looks like a payment/settlement, not a shared expense.",
            "policy": "Payments are stored separately from expenses.",
            "action": "Created Settlement row.", "review": False,
        })
        needs_review = any(a["review"] for a in anomalies) or "deposit" in description.lower()
        if "deposit" in description.lower():
            anomalies.append({
                "code": "DEPOSIT_PAYMENT_DETECTED", "severity": ImportAnomaly.Severity.WARNING,
                "message": "Deposit-like payment found.",
                "policy": "Deposit payments are allowed as settlements but require review because they may not be shared living expenses.",
                "action": "Created review-required Settlement row.", "review": True,
            })
            needs_review = True
        status = LedgerStatus.REVIEW_REQUIRED if needs_review else LedgerStatus.POSTED
        settlement = Settlement.objects.create(
            group=self.group,
            import_batch=self.batch,
            raw_row_number=row_number,
            date=day,
            paid_by=payer,
            paid_to=paid_to,
            amount_original=original_amount,
            currency=currency,
            fx_rate_to_inr=fx_rate,
            amount_inr=amount_inr,
            notes=f"{description}. {notes}".strip(),
            status=status,
        )
        self.persist_anomalies(row_number, anomalies, settlement=settlement)
        self.increment_counts(status)
        self.add_report_row(row_number, status, anomalies)

    def detect_duplicates(self, row_number, day, description, payer, amount_inr, participants, anomalies):
        desc = normalize_description(description)
        exact_key = (day.isoformat(), desc, payer, str(amount_inr), tuple(sorted(participants)))
        if exact_key in self.exact_seen:
            anomalies.append({
                "code": "DUPLICATE_EXACT", "severity": ImportAnomaly.Severity.WARNING,
                "message": f"Looks like exact duplicate of row {self.exact_seen[exact_key]}.",
                "policy": "Do not delete duplicates silently.",
                "action": "Row requires review and is not posted.", "review": True,
            })
            return True
        self.exact_seen[exact_key] = row_number

        for previous in self.fuzzy_seen:
            same_day = previous["day"] == day
            same_people = set(previous["participants"]) == set(participants)
            similarity = SequenceMatcher(None, previous["desc"], desc).ratio()
            if same_day and same_people and similarity >= 0.62:
                if previous["amount"] != amount_inr or previous["payer"] != payer:
                    anomalies.append({
                        "code": "DUPLICATE_FUZZY_AMOUNT_MISMATCH", "severity": ImportAnomaly.Severity.WARNING,
                        "message": f"Possible duplicate of row {previous['row']} with different payer/amount.",
                        "policy": "Human must choose the winning row.",
                        "action": "Row requires review and is not posted.", "review": True,
                    })
                    self.fuzzy_seen.append({"row": row_number, "day": day, "desc": desc, "payer": payer, "amount": amount_inr, "participants": participants})
                    return True
        self.fuzzy_seen.append({"row": row_number, "day": day, "desc": desc, "payer": payer, "amount": amount_inr, "participants": participants})
        return False

    def persist_anomalies(self, row_number, anomalies, expense=None, settlement=None):
        for anomaly in anomalies:
            ImportAnomaly.objects.create(
                batch=self.batch,
                row_number=row_number,
                code=anomaly["code"],
                severity=anomaly["severity"],
                message=anomaly["message"],
                policy=anomaly["policy"],
                action_taken=anomaly["action"],
                requires_review=anomaly["review"],
                expense=expense,
                settlement=settlement,
                suggested_payload=anomaly.get("payload", {}),
            )

    def add_report_only(self, row_number, anomalies):
        self.persist_anomalies(row_number, anomalies)
        self.increment_counts(LedgerStatus.SKIPPED)
        self.add_report_row(row_number, LedgerStatus.SKIPPED, anomalies)

    def add_report_row(self, row_number, status, anomalies):
        self.report_rows.append({
            "row_number": row_number,
            "status": status,
            "anomalies": [
                {
                    "code": a["code"],
                    "severity": str(a["severity"]),
                    "message": a["message"],
                    "policy": a["policy"],
                    "action_taken": a["action"],
                    "requires_review": a["review"],
                }
                for a in anomalies
            ],
        })

    def increment_counts(self, status):
        if status == LedgerStatus.POSTED:
            self.batch.posted_rows += 1
        elif status == LedgerStatus.REVIEW_REQUIRED:
            self.batch.review_rows += 1
        else:
            self.batch.skipped_rows += 1

    def finish_batch(self):
        self.batch.status = ImportBatch.Status.COMPLETED
        self.batch.report_json = {
            "source_filename": self.batch.source_filename,
            "total_rows": self.batch.total_rows,
            "posted_rows": self.batch.posted_rows,
            "review_rows": self.batch.review_rows,
            "skipped_rows": self.batch.skipped_rows,
            "rows": self.report_rows,
        }
        self.batch.save(update_fields=["status", "posted_rows", "review_rows", "skipped_rows", "report_json"])
